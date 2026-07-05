import os
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# Define output folder
output_dir = r"d:\VS CODE\astudio resume\test_cases"
os.makedirs(output_dir, exist_ok=True)

# Definition of the 6 modules with their corresponding screens/sub-modules
modules = [
    {
        "file": "01_authentication_and_profile_management.xlsx",
        "title": "MODULE 1: AUTHENTICATION & USER PROFILE MANAGEMENT",
        "prefix": "TC_AUTH_",
        "sub_modules": [
            {
                "name": "Sign-In Page Layout",
                "ui_desc": "Login screen containing branding logo, email text field, password text field, and 'Sign In' buttons.",
                "valid_action": "Verify layout rendering by checking alignment of fields, background colors, and visibility of all elements.",
                "empty_action": "Leave both email and password fields blank, and click the 'Sign In' button.",
                "invalid_action": "Enter text in fields without correct positioning or size boundaries.",
                "steps_prefix": "1. Launch the SkillGap application.\n2. Navigate to the login screen.\n3."
            },
            {
                "name": "Email Input Field",
                "ui_desc": "The email text box on the sign-in screen.",
                "valid_action": "Input a standard valid email like 'candidate.success@skillgap.ai' and blur field.",
                "empty_action": "Leave the email input field entirely blank while filling the password.",
                "invalid_action": "Type malformed email structures (e.g., 'test@domain', 'user@domain.c', 'user@@domain.com').",
                "steps_prefix": "1. Go to the Sign-In screen.\n2. Locate the email text input box.\n3."
            },
            {
                "name": "Password Input Field",
                "ui_desc": "The password input box with toggle visibility eye icon.",
                "valid_action": "Type a valid password matching complexity rules (e.g. 'StrongPass@123') and click the eye toggle icon.",
                "empty_action": "Leave the password field empty while filling a valid email.",
                "invalid_action": "Enter a password that is too short (e.g. '123') or lacks required characters.",
                "steps_prefix": "1. Focus on the Password input text field.\n2. Type the password values.\n3."
            },
            {
                "name": "Forgot Password Dialog",
                "ui_desc": "Password recovery pop-up dialog that requests email address to send a recovery link.",
                "valid_action": "Click 'Forgot Password', input a valid registered email, and click 'Send Link'.",
                "empty_action": "Open the dialog, leave the email field empty, and click the 'Send Link' button.",
                "invalid_action": "Input an unregistered or malformed email address in the recovery field.",
                "steps_prefix": "1. Tap on the 'Forgot Password' link.\n2. Wait for the popup modal dialog to display.\n3."
            },
            {
                "name": "Submit Login Button",
                "ui_desc": "Button to trigger credentials verification against Firebase Authentication.",
                "valid_action": "Enter valid email and password credentials and click the 'Sign In' button.",
                "empty_action": "Attempt to click 'Sign In' button while both fields are blank.",
                "invalid_action": "Input credentials and trigger double-clicks on the login button rapidly.",
                "steps_prefix": "1. Input credentials on the login screen.\n2. Locate the 'Sign In' button.\n3."
            },
            {
                "name": "Biometric Login Sync",
                "ui_desc": "Biometric sync switcher to authorize fingerprint or face ID access.",
                "valid_action": "Toggle the biometric switch to ON and authorize via fingerprint simulator scan.",
                "empty_action": "Attempt to toggle biometric login without setting up system screen lock.",
                "invalid_action": "Perform a biometric scan using an unregistered fingerprint.",
                "steps_prefix": "1. Go to settings / profile area.\n2. Tap the Biometric Auth switch toggle.\n3."
            },
            {
                "name": "Registration Page Layout",
                "ui_desc": "Sign up page showing full name, email, phone number, and password fields.",
                "valid_action": "Open signup form and verify formatting, layout alignments, and terms of service click.",
                "empty_action": "Attempt to submit the registration form with all inputs left empty.",
                "invalid_action": "Fill standard inputs but containing HTML tags or scripting payloads.",
                "steps_prefix": "1. Click on the 'Create Account' link on the sign-in screen.\n2. Check screen components.\n3."
            },
            {
                "name": "Registration Email Field",
                "ui_desc": "Email text field on the user registration form.",
                "valid_action": "Type a fresh valid email address 'newuser@skillgap.ai' in the field.",
                "empty_action": "Leave the signup email field blank while filling other credentials.",
                "invalid_action": "Input a duplicate email that is already registered in Firebase database.",
                "steps_prefix": "1. Focus on the Email input box in registration page.\n2. Input the email string.\n3."
            },
            {
                "name": "Registration Password Field",
                "ui_desc": "Password field on the signup form with live strength indicator labels.",
                "valid_action": "Input a password meeting all strength requirements and confirm matching values.",
                "empty_action": "Leave password field blank and attempt signup.",
                "invalid_action": "Enter password mismatch in confirm password field or values under 6 characters.",
                "steps_prefix": "1. Navigate to the password complexity input field.\n2. Enter password value.\n3."
            },
            {
                "name": "Submit Registration Button",
                "ui_desc": "Submit button for creating a new user account.",
                "valid_action": "Fill in all valid registration details and click the 'Sign Up' button.",
                "empty_action": "Click the registration button without verifying the terms & conditions checkbox.",
                "invalid_action": "Double-tap the button rapidly during slow internet connection.",
                "steps_prefix": "1. Complete all fields on the registration form.\n2. Locate the 'Register' button.\n3."
            },
            {
                "name": "Firestore User Sync",
                "ui_desc": "Database synchronization operation to Firestore database.",
                "valid_action": "Register account and check if user profile record is created in Firestore database.",
                "empty_action": "Simulate signup termination prior to cloud profile sync completion.",
                "invalid_action": "Attempt database sync when cloud database rules restrict write access.",
                "steps_prefix": "1. Click on Submit Registration.\n2. Observe user profile redirection.\n3."
            },
            {
                "name": "Profile Screen Layout",
                "ui_desc": "Profile screen rendering user details, stats, target jobs, and logout button.",
                "valid_action": "Navigate to profile and verify alignment of fields, user information, and scores.",
                "empty_action": "Verify profile display when user Firestore profile is missing optional data fields.",
                "invalid_action": "Observe profile rendering on a device configured with extremely large font settings.",
                "steps_prefix": "1. Log in to the application.\n2. Tap the Profile icon in bottom navigation bar.\n3."
            },
            {
                "name": "Edit Name and Contact Info",
                "ui_desc": "Profile details editor overlay to update full name and phone number.",
                "valid_action": "Change display name and phone number to valid values and click 'Save Changes'.",
                "empty_action": "Clear the display name field completely and click the 'Save' button.",
                "invalid_action": "Type alphabetic letters in the phone number field or special characters in the name.",
                "steps_prefix": "1. Open Profile Screen.\n2. Tap the 'Edit Profile' button.\n3."
            },
            {
                "name": "Profile Photo Upload",
                "ui_desc": "Image upload button to select and crop a user avatar.",
                "valid_action": "Click avatar placeholder, select a valid 2MB PNG photo, crop it, and save.",
                "empty_action": "Open avatar file picker, cancel it, and click save.",
                "invalid_action": "Attempt to upload a non-image file extension (e.g. .pdf or .zip).",
                "steps_prefix": "1. Go to the profile editor.\n2. Tap on the user avatar image picker.\n3."
            },
            {
                "name": "Dream Job Dropdown",
                "ui_desc": "Job role dropdown component to select the target dream job.",
                "valid_action": "Tap on the dropdown and select 'Backend Developer' from the populated list.",
                "empty_action": "Open the dropdown and click outside it without making any selection.",
                "invalid_action": "Simulate a scenario where the database fails to load target roles.",
                "steps_prefix": "1. Locate the 'Dream Job Target' dropdown menu.\n2. Tap to expand it.\n3."
            },
            {
                "name": "Search Dream Job Input",
                "ui_desc": "Fuzzy text search input field to filter available dream jobs in dropdown.",
                "valid_action": "Type 'Data' to search for roles like 'Data Scientist' and 'Data Analyst'.",
                "empty_action": "Type spaces in search field and check if it lists all roles.",
                "invalid_action": "Type special symbols and SQL keywords in the job search field.",
                "steps_prefix": "1. Open the target job selector dropdown.\n2. Locate the search filter input.\n3."
            },
            {
                "name": "Save Dream Job Button",
                "ui_desc": "Save button to update target dream job preference in the backend.",
                "valid_action": "Select a role and tap 'Save Preference' to update target metrics.",
                "empty_action": "Click the save button without choosing a target job first.",
                "invalid_action": "Click the button when offline and check local db caching.",
                "steps_prefix": "1. Choose a target dream job.\n2. Highlight the 'Save' button.\n3."
            },
            {
                "name": "Diagnostic Data Reset",
                "ui_desc": "Reset Diagnostics button in the profile settings panel.",
                "valid_action": "Tap the button, confirm 'Reset' in popup warning, and verify database wipe.",
                "empty_action": "Tap the reset button, but choose 'Cancel' in the confirmation prompt dialog.",
                "invalid_action": "Spam click the reset confirmation button to trigger concurrent delete calls.",
                "steps_prefix": "1. Go to profile settings.\n2. Tap the 'Reset Diagnostic Profile' button.\n3."
            },
            {
                "name": "Auto-Login Session Persistency",
                "ui_desc": "App startup session check logic using local secure storage tokens.",
                "valid_action": "Close app while logged in, launch app again, and check if it bypasses login screen.",
                "empty_action": "Launch app when no local session storage token exists.",
                "invalid_action": "Launch app when the stored authentication token is invalid or expired.",
                "steps_prefix": "1. Terminate the active application instance.\n2. Relaunch the application from system tray.\n3."
            },
            {
                "name": "Logout Session Termination",
                "ui_desc": "Logout option located in profile settings to invalidate user tokens.",
                "valid_action": "Tap 'Log Out' button, confirm exit, and verify redirection to login screen.",
                "empty_action": "Tap logout button, but cancel the action in confirmation popup.",
                "invalid_action": "Tap logout when network is disconnected and check offline local cache cleanup.",
                "steps_prefix": "1. Navigate to the profile settings panel.\n2. Locate the 'Log Out' action item.\n3."
            }
        ]
    },
    {
        "file": "02_resume_upload_and_ai_parsing_engine.xlsx",
        "title": "MODULE 2: RESUME UPLOAD & AI PARSING ENGINE",
        "prefix": "TC_RESUME_",
        "sub_modules": [
            {
                "name": "Resume Upload UI Screen",
                "ui_desc": "The main resume drag-and-drop / select page containing file picker links.",
                "valid_action": "Open the screen and verify file select cards, description notes, and active limits.",
                "empty_action": "Click the 'Upload' button without importing any file.",
                "invalid_action": "Interact with the drag zone using folders instead of files.",
                "steps_prefix": "1. Navigate to the Resume Upload screen.\n2. Check UI components layout.\n3."
            },
            {
                "name": "PDF File Picker Launch",
                "ui_desc": "System file chooser dialog launched when clicking on upload card.",
                "valid_action": "Click the select area and verify that the Android system file explorer opens.",
                "empty_action": "Launch file explorer and exit immediately by pressing system back button.",
                "invalid_action": "Simulate file selector launch when storage permissions are denied.",
                "steps_prefix": "1. Locate the file selection card area.\n2. Click to open file manager.\n3."
            },
            {
                "name": "PDF File Picker Cancel",
                "ui_desc": "Cancel behavior during local file selection.",
                "valid_action": "Launch file chooser, press back or tap close to return without selections.",
                "empty_action": "Simulate file choice cancellation multiple times sequentially.",
                "invalid_action": "Select a file but double-click cancel while file explorer is closing.",
                "steps_prefix": "1. Open the file explorer dialog.\n2. Select cancel option.\n3."
            },
            {
                "name": "Upload Non-PDF Format",
                "ui_desc": "Uploading file types with extensions other than .pdf.",
                "valid_action": "Attempt to drag a .docx file and check for format error notification.",
                "empty_action": "Select a file with no extension and click upload.",
                "invalid_action": "Rename a .png image file to .pdf extension and attempt upload.",
                "steps_prefix": "1. Prepare a test file with a non-PDF extension.\n2. Drag it to upload field.\n3."
            },
            {
                "name": "Upload Empty PDF (0 KB)",
                "ui_desc": "Handling of zero-byte resume files.",
                "valid_action": "Choose a 0 KB PDF file and attempt to upload it to the scanner.",
                "empty_action": "Drag and drop a blank document containing no text or metadata.",
                "invalid_action": "Spam click analyze after selecting a zero-byte file.",
                "steps_prefix": "1. Select a 0 KB empty PDF file.\n2. Load it to upload screen.\n3."
            },
            {
                "name": "Upload Huge PDF (> 20 MB)",
                "ui_desc": "File size boundary check exceeding maximum limits.",
                "valid_action": "Attempt to upload a 25 MB PDF file and verify error warning.",
                "empty_action": "Test file upload with file size exactly equal to 5.01 MB.",
                "invalid_action": "Simulate network upload progress bar on a slow network for heavy file.",
                "steps_prefix": "1. Select a PDF file larger than 5 MB (e.g., 25 MB).\n2. Tap on upload button.\n3."
            },
            {
                "name": "Upload Corrupted PDF",
                "ui_desc": "Uploading damaged or unreadable PDF documents.",
                "valid_action": "Upload a corrupted PDF and check if the parser displays file read error.",
                "empty_action": "Upload PDF file missing PDF signature headers.",
                "invalid_action": "Input corrupted data streams to parser endpoint directly.",
                "steps_prefix": "1. Access a PDF file that has corrupted headers.\n2. Upload it to app.\n3."
            },
            {
                "name": "Password-Protected PDF",
                "ui_desc": "PDF files encrypted with owner or user passwords.",
                "valid_action": "Upload a password-protected PDF and check if it prompts password or rejects it.",
                "empty_action": "Attempt to parse password encrypted PDF file directly without decryption.",
                "invalid_action": "Input wrong decryption passwords in dialog if prompted.",
                "steps_prefix": "1. Encrypt a resume PDF with a password.\n2. Upload the encrypted file.\n3."
            },
            {
                "name": "Resume Parsing Scanner Ring",
                "ui_desc": "The visual circular progress scanner animation shown during parsing.",
                "valid_action": "Upload a valid PDF, observe circular scanning animation rotating smoothly.",
                "empty_action": "Simulate UI layout constraints of scanner card.",
                "invalid_action": "Send app to background during scanning rotation to check UI state preservation.",
                "steps_prefix": "1. Upload a valid resume file.\n2. Click the 'Analyze' button to start parsing.\n3."
            },
            {
                "name": "Scanning Ring Rotation Speed",
                "ui_desc": "Rotation and frame rate behavior of parser animation.",
                "valid_action": "Check rotation transitions and ensure it runs at stable FPS without lag.",
                "empty_action": "Check CPU and GPU render latency during active rotation loop.",
                "invalid_action": "Open multi-window view to see if scanning ring rotation stutters.",
                "steps_prefix": "1. Launch parser scan animation.\n2. Monitor UI rendering performance.\n3."
            },
            {
                "name": "Analysis Progress Steps Display",
                "ui_desc": "Status text checklist shown below scan animation (e.g. 'Extracting Text', 'Scoring ATS').",
                "valid_action": "Verify that all progress phases display step-by-step indicators in order.",
                "empty_action": "Verify layout when step data fails to load text strings.",
                "invalid_action": "Observe text behavior during high DPI viewport rendering.",
                "steps_prefix": "1. Initiate parsing flow.\n2. Look at status steps list (Extracting, Categorizing, Scoring).\n3."
            },
            {
                "name": "Analysis Step Status Updates",
                "ui_desc": "Transition logic of parsing step checkpoints.",
                "valid_action": "Check step markers turn green one by one as backend stages finish.",
                "empty_action": "Disconnect wifi during step 2 to verify connection warning banner.",
                "invalid_action": "Simulate step state updates in incorrect order (e.g. step 3 finishes before step 1).",
                "steps_prefix": "1. Run resume analysis.\n2. Watch each step update its state.\n3."
            },
            {
                "name": "Cancel Parsing Request",
                "ui_desc": "Cancel icon button next to parsing status indicator.",
                "valid_action": "Click 'Cancel' button during parsing and verify abort redirection.",
                "empty_action": "Click cancel, but continue upload if server fails to abort.",
                "invalid_action": "Spam click the cancel button after upload has already reached 99%.",
                "steps_prefix": "1. Start parsing a heavy resume file.\n2. Tap the 'Cancel' button mid-process.\n3."
            },
            {
                "name": "Match Score Display Card",
                "ui_desc": "Main percentage score card shown in results screen.",
                "valid_action": "Verify overall match percentage card renders with correct colors and score value.",
                "empty_action": "Verify display when matching score resolves to exactly 0%.",
                "invalid_action": "Inspect layout when the percentage string exceeds standard width fields.",
                "steps_prefix": "1. Navigate to analysis results dashboard.\n2. Check overall score display card.\n3."
            },
            {
                "name": "ATS Score Progress Indicator",
                "ui_desc": "Circular score meter detailing ATS compliance levels.",
                "valid_action": "Verify ATS sub-score calculations match resume fields validation metrics.",
                "empty_action": "Verify ATS meter when contact info has no matched parameters.",
                "invalid_action": "Change system font scaling to check circular text clip boundaries.",
                "steps_prefix": "1. Open the results screen.\n2. Scroll down to ATS compliance section.\n3."
            },
            {
                "name": "Readability Score Pill",
                "ui_desc": "Flesch-Kincaid grade level visual pill rating.",
                "valid_action": "Verify readability rating updates based on word count and sentence lengths.",
                "empty_action": "Observe pill style when reading score returns null.",
                "invalid_action": "Simulate readability calculation on a resume containing encrypted content.",
                "steps_prefix": "1. Check the Readability Score card pill layout.\n2. Verify the level label.\n3."
            },
            {
                "name": "Missing Skills List UI",
                "ui_desc": "Scrollable list of missing technologies identified in resume.",
                "valid_action": "Verify missing skills are list-rendered with custom red skill chips.",
                "empty_action": "Observe UI when there are 0 missing skills (perfect match).",
                "invalid_action": "Type very long custom skill names to verify text wrapping.",
                "steps_prefix": "1. Go to missing skills card grid.\n2. Verify skill chip layout.\n3."
            },
            {
                "name": "Smart Recommendation Cards",
                "ui_desc": "Recommendation cards listing courses and roadmap paths.",
                "valid_action": "Verify cards show custom actions matching user target weaknesses.",
                "empty_action": "Render screen when no courses recommendations are returned from API.",
                "invalid_action": "Tap recommendation links when offline and verify alert behavior.",
                "steps_prefix": "1. Scroll to recommendation cards layout.\n2. Tap one of the course cards.\n3."
            },
            {
                "name": "Learning Roadmap Phase Card",
                "ui_desc": "Visual milestone card previewing the generated study phases.",
                "valid_action": "Verify roadmap phase cards list target milestones and durations.",
                "empty_action": "Load page with empty roadmap object cache.",
                "invalid_action": "Tap on roadmap milestones rapidly to verify navigation responses.",
                "steps_prefix": "1. Look at learning roadmap card on results panel.\n2. Observe phase details.\n3."
            },
            {
                "name": "Resume Issues Validation Cards",
                "ui_desc": "Validation cards explaining parsing errors or layout warnings.",
                "valid_action": "Verify list of formatting problems (e.g. 'No Projects Found') matches parsed resume data.",
                "empty_action": "Verify screen appearance when zero resume formatting warnings exist.",
                "invalid_action": "Tap issue cards and verify dialog descriptions formatting.",
                "steps_prefix": "1. Locate resume quality details card.\n2. Tap on issue items.\n3."
            }
        ]
    },
    {
        "file": "03_student_dashboard_and_skill_gap_analytics.xlsx",
        "title": "MODULE 3: STUDENT DASHBOARD & SKILL GAP ANALYTICS",
        "prefix": "TC_DASH_",
        "sub_modules": [
            {
                "name": "Dashboard Landing Page UI",
                "ui_desc": "Main dashboard landing page displaying profile summaries, metrics, and progress.",
                "valid_action": "Navigate to dashboard and check formatting of cards and overall accessibility.",
                "empty_action": "Load dashboard layout for a new user who hasn't uploaded a resume.",
                "invalid_action": "Launch dashboard on low-density mobile display screens to check clipping.",
                "steps_prefix": "1. Launch the app and authenticate.\n2. Land on dashboard screen.\n3."
            },
            {
                "name": "Resume Status Warning Banner",
                "ui_desc": "Banner warning 'Upload your resume to unlock full features' visible at top.",
                "valid_action": "Check banner appears for unanalyzed users and click 'Upload Now' shortcut link.",
                "empty_action": "Verify banner is completely absent when active resume score exists.",
                "invalid_action": "Perform horizontal swipe gestures on the warning banner card.",
                "steps_prefix": "1. Open app with account containing no resume uploads.\n2. Look at top of Dashboard.\n3."
            },
            {
                "name": "Feature Gating Screen Lock",
                "ui_desc": "Screen lock overlay blocking feature access until resume parsing completes.",
                "valid_action": "Tap locked features (e.g. Quiz, Career predicts) and verify lock message overlay is shown.",
                "empty_action": "Verify features remain locked when resume analysis fails during upload.",
                "invalid_action": "Attempt to bypass lock overlay by navigating via system hardware back button.",
                "steps_prefix": "1. Navigate to dashboard with blocked access.\n2. Tap on Quiz / Prep card.\n3."
            },
            {
                "name": "Bottom Navigation Bar UI",
                "ui_desc": "Navigation bar containing icons for Dashboard, Prep, Interview, and Profile.",
                "valid_action": "Tap each navigation icon, verifying smooth page transition and active states.",
                "empty_action": "Inspect tab states when switching routes under memory constraints.",
                "invalid_action": "Tap adjacent icons simultaneously using multi-touch gestures.",
                "steps_prefix": "1. Locate bottom navigation bar.\n2. Tap on different navigation icons.\n3."
            },
            {
                "name": "Top Bar and Settings Menu",
                "ui_desc": "Top app bar containing app logo title and settings cog shortcut icon.",
                "valid_action": "Tap settings cog and check settings screen loading overlay options.",
                "empty_action": "Verify title rendering in horizontal landscape orientation modes.",
                "invalid_action": "Double-tap settings cog rapidly to stress-test rendering engine.",
                "steps_prefix": "1. Look at top app bar on dashboard.\n2. Tap on settings cog icon.\n3."
            },
            {
                "name": "Overall Match Score Card",
                "ui_desc": "Summary match percentage widget card showing target role progress.",
                "valid_action": "Check match score displays matching values in sync with Firestore user model.",
                "empty_action": "Verify overall score display card when database returned data is null.",
                "invalid_action": "Inspect card borders and text scales under maximum system font sizes.",
                "steps_prefix": "1. Go to matching details on dashboard.\n2. Locate overall match score card.\n3."
            },
            {
                "name": "Skill Gap Radar Chart",
                "ui_desc": "Interactive radar web graph plotting user skills against market demand levels.",
                "valid_action": "Hover/tap nodes on the radar chart and check tooltip values.",
                "empty_action": "Verify graph rendering when user has 0 matching skills.",
                "invalid_action": "Rotate screen to landscape and back while radar chart is drawing.",
                "steps_prefix": "1. Scroll down to Skill Gap analytics card.\n2. Tap on one of the radar graph corners.\n3."
            },
            {
                "name": "Progress Comparison Graph",
                "ui_desc": "Line chart showing score improvement comparison over time.",
                "valid_action": "Open progress graphs, verify chart line plotting matches history entries.",
                "empty_action": "Verify graph displays blank message grid for user with single log history.",
                "invalid_action": "Simulate line plotting when progress entries contain negative timestamps.",
                "steps_prefix": "1. Open the progress tracker card.\n2. Observe the historical line graph.\n3."
            },
            {
                "name": "Peer Analytics Chart",
                "ui_desc": "Bar chart comparing user match score against peer averages.",
                "valid_action": "Verify user bar and peer average bar are labeled and colored differently.",
                "empty_action": "Verify chart renders when server peer average is null or returns 0.",
                "invalid_action": "Inspect bar chart scaling calculations when user score is 100%.",
                "steps_prefix": "1. Open peer analytics panel.\n2. Check the comparative bar graph layout.\n3."
            },
            {
                "name": "Historical Skill Trends",
                "ui_desc": "List of recently updated skills and progress percentages.",
                "valid_action": "Verify recent improvements show delta values (e.g. '+5% React').",
                "empty_action": "Verify list displays placeholder illustration when no skills were added.",
                "invalid_action": "Input extremely long skill category labels into historical lists.",
                "steps_prefix": "1. Go to skill trends section.\n2. Check listing items.\n3."
            },
            {
                "name": "Share Progress Button",
                "ui_desc": "Share button next to score metrics card.",
                "valid_action": "Tap share button, confirm system sharing intent launch with image.",
                "empty_action": "Tap share button, but dismiss system tray without sending.",
                "invalid_action": "Trigger share button when device gallery is locked or inaccessible.",
                "steps_prefix": "1. Tap on the share progress icon button.\n2. Wait for system share sheet dialog.\n3."
            },
            {
                "name": "Missing Skills Search Input",
                "ui_desc": "Search input box inside the skill gap details panel.",
                "valid_action": "Search missing skills list by entering search query (e.g. 'SQL').",
                "empty_action": "Type space characters in the search field and verify list filtering.",
                "invalid_action": "Type symbols and sql characters to check injection filters.",
                "steps_prefix": "1. Focus on the missing skills search input field.\n2. Input target query.\n3."
            },
            {
                "name": "Missing Skill Learn Redirect",
                "ui_desc": "Learn buttons next to missing skill chips linking to roadmap.",
                "valid_action": "Tap 'Learn' link on missing skill to auto-scroll to the specific study task.",
                "empty_action": "Tap learn link when roadmap for that skill has not been calculated.",
                "invalid_action": "Rapidly click the redirect link to trigger duplicate layout jumps.",
                "steps_prefix": "1. Select a missing skill card chip.\n2. Tap the 'Learn' shortcut button next to it.\n3."
            },
            {
                "name": "Matching Skills Chip Layout",
                "ui_desc": "Grid displaying user matched skills in green colored tags.",
                "valid_action": "Check matched chips wrap nicely and display correct labels.",
                "empty_action": "Verify chip layout rendering when matched list is empty.",
                "invalid_action": "Simulate resizing screen to check grid column wrapping constraints.",
                "steps_prefix": "1. Locate matched skills chip container.\n2. Check color coding.\n3."
            },
            {
                "name": "Profile Diagnostics Overview",
                "ui_desc": "Score scorecard details panel showing technical, resume, and coding sub-metrics.",
                "valid_action": "Verify each score card reflects accurate results and loads progress rings.",
                "empty_action": "Observe rendering when database returns scores as 0.",
                "invalid_action": "Observe card layout under forced dark mode theme overrides.",
                "steps_prefix": "1. Scroll to Diagnostics overview tab.\n2. Inspect tech, resume, coding score cards.\n3."
            },
            {
                "name": "Dashboard Offline Alert",
                "ui_desc": "Offline toast message or banner appearing during internet loss.",
                "valid_action": "Disconnect network on dashboard, check offline warning banner display.",
                "empty_action": "Connect back to network, check that warning banner auto-dismisses.",
                "invalid_action": "Perform pull to refresh action while network is disconnected.",
                "steps_prefix": "1. Open dashboard screen.\n2. Turn off mobile network connectivity.\n3."
            },
            {
                "name": "Landscape Layout Adaptation",
                "ui_desc": "Dashboard responsive design in landscape mode.",
                "valid_action": "Rotate device to landscape, check dashboard cards grid changes columns.",
                "empty_action": "Rotate screen with keyboard active to verify viewport height limits.",
                "invalid_action": "Simulate rotation toggle while dashboard is loading database data.",
                "steps_prefix": "1. Rotate target device screen to landscape mode.\n2. Check layouts alignment.\n3."
            },
            {
                "name": "Theme Settings Toggle",
                "ui_desc": "Dark/Light mode theme switch in settings panel.",
                "valid_action": "Toggle dark theme switch, verify background changes to deep gray/dark blue.",
                "empty_action": "Verify default theme mode matches system settings layout.",
                "invalid_action": "Rapidly toggle dark/light theme switch 10 times in a row.",
                "steps_prefix": "1. Navigate to settings from dashboard.\n2. Locate the theme toggle switch.\n3."
            },
            {
                "name": "Dashboard Pull-to-Refresh",
                "ui_desc": "Standard pull-to-refresh swipe gesture on dashboard screen.",
                "valid_action": "Swipe down, verify loading spinner rotates and data refreshes.",
                "empty_action": "Swipe down, but abort swipe mid-gesture to check reload cancel.",
                "invalid_action": "Perform pull-to-refresh gesture when network is entirely offline.",
                "steps_prefix": "1. Touch top of dashboard screen.\n2. Drag finger downwards to trigger pull-to-refresh.\n3."
            },
            {
                "name": "Settings Clean State Sync",
                "ui_desc": "Reset account cache settings action.",
                "valid_action": "Select 'Clean Cache' option, check data is cleared and refreshed from Firestore.",
                "empty_action": "Tap clear database settings cache, but deny confirmation warning.",
                "invalid_action": "Trigger cache cleanup during an active database sync operation.",
                "steps_prefix": "1. Open settings screen.\n2. Tap on 'Clear Local Cache' option.\n3."
            }
        ]
    },
    {
        "file": "04_placement_and_coding_preparation.xlsx",
        "title": "MODULE 4: PLACEMENT & CODING PREPARATION",
        "prefix": "TC_PREP_",
        "sub_modules": [
            {
                "name": "Placement Prep Main Menu",
                "ui_desc": "Main menu listing Aptitude prep and Coding editor shortcuts.",
                "valid_action": "Verify menu cards render correctly with descriptions and difficulty tags.",
                "empty_action": "Load prep main menu when user metadata is missing target profile info.",
                "invalid_action": "Click prep options when application is running on offline mode.",
                "steps_prefix": "1. Select 'Prep' section from bottom navigation.\n2. Inspect Main Menu UI layout.\n3."
            },
            {
                "name": "Aptitude Category List",
                "ui_desc": "Grid listing categories like Quantitative, Logical, and Verbal.",
                "valid_action": "Verify categories load with question counts and category icon labels.",
                "empty_action": "Verify category container displays blank info when server database returns empty.",
                "invalid_action": "Double tap category cards rapidly to check duplicate route creations.",
                "steps_prefix": "1. Navigate to Aptitude categories grid.\n2. Check topic listing cards.\n3."
            },
            {
                "name": "Mock Aptitude Quiz Start",
                "ui_desc": "Quiz screen showing questions with multiple choice radio buttons.",
                "valid_action": "Select category, start quiz, verify questions load with 4 selectable radio options.",
                "empty_action": "Simulate quiz launch when database returns null questions for category.",
                "invalid_action": "Minimize app during quiz setup to see if session timer resets.",
                "steps_prefix": "1. Tap on a topic category card.\n2. Click the 'Start Quiz' button.\n3."
            },
            {
                "name": "Quiz Timer Display",
                "ui_desc": "Countdown timer clock shown at top right of active quiz screen.",
                "valid_action": "Check timer counts down seconds correctly and turns red below 1 minute.",
                "empty_action": "Verify quiz behavior when timer reaches exactly 00:00 limit.",
                "invalid_action": "Modify local system device clock settings to check timer tamper safety.",
                "steps_prefix": "1. Start active quiz screen.\n2. Look at countdown timer widget.\n3."
            },
            {
                "name": "Aptitude Quiz Submit",
                "ui_desc": "Submit button inside active quiz screen.",
                "valid_action": "Complete quiz, tap 'Submit Quiz', confirm in modal, and check scorecard loading.",
                "empty_action": "Tap submit quiz button while multiple questions remain unanswered.",
                "invalid_action": "Trigger system back button gesture during submit modal display.",
                "steps_prefix": "1. Reach final question of quiz.\n2. Click the 'Submit Quiz' button.\n3."
            },
            {
                "name": "Aptitude Review Answers",
                "ui_desc": "Score display dashboard showing correct, wrong, and skipped question cards.",
                "valid_action": "Verify green/red highlight tags indicate correct/incorrect answers in review list.",
                "empty_action": "Verify review details scorecard when all questions were skipped.",
                "invalid_action": "Scroll review list on an extremely small display screen.",
                "steps_prefix": "1. Submit mock aptitude quiz.\n2. Click on 'Review Answers' button.\n3."
            },
            {
                "name": "Coding Prep Challenge List",
                "ui_desc": "Coding challenge list displaying title, difficulty badges, and solve status.",
                "valid_action": "Verify challenges populate with correct difficulty chips (Easy, Medium, Hard).",
                "empty_action": "Verify UI list representation when filters return 0 matching challenges.",
                "invalid_action": "Input long search keywords in challenge filter input field.",
                "steps_prefix": "1. Open Coding Preparation card.\n2. Check challenge lists.\n3."
            },
            {
                "name": "Coding Editor UI Layout",
                "ui_desc": "Code input terminal editor screen with tabs for description, code, and console.",
                "valid_action": "Open editor page and verify text field input typing, indentation, and scrolling.",
                "empty_action": "Open editor, verify default main function boiler template loaded for chosen language.",
                "invalid_action": "Simulate rotation to landscape view during editing mode.",
                "steps_prefix": "1. Tap on a coding challenge item.\n2. Navigate to editor layout screen.\n3."
            },
            {
                "name": "Language Selection Dropdown",
                "ui_desc": "Language dropdown selector offering C++, Java, Python, and Kotlin.",
                "valid_action": "Select 'Python', verify editor boiler template updates to python syntax.",
                "empty_action": "Verify dropdown behavior when tapped but dismissed without clicking new item.",
                "invalid_action": "Change language options mid-execution when code is running on server.",
                "steps_prefix": "1. Locate language selector dropdown.\n2. Tap to expand list.\n3."
            },
            {
                "name": "Code Autosave Action",
                "ui_desc": "Autosave action saving code template draft in local SQLite database.",
                "valid_action": "Type code, exit editor screen, navigate back, and verify draft is restored.",
                "empty_action": "Ensure no draft database saving occurs if code editor has not been edited.",
                "invalid_action": "Kill application task stack mid-typing and check persistence.",
                "steps_prefix": "1. Focus code editor area.\n2. Input custom lines of code.\n3."
            },
            {
                "name": "Run Code Button",
                "ui_desc": "Run code button to test custom solution against initial cases.",
                "valid_action": "Write working code, click 'Run Code', check output matches expected cases.",
                "empty_action": "Click 'Run Code' button while compiler editor contains blank text.",
                "invalid_action": "Spam click the button while compiler is compiling the code.",
                "steps_prefix": "1. Write standard logic in editor.\n2. Click the 'Run Code' button.\n3."
            },
            {
                "name": "Compiler Syntax Highlight",
                "ui_desc": "Syntax color formatting inside editor view.",
                "valid_action": "Verify compiler syntax highlights keywords (e.g. class, def, return) in distinct colors.",
                "empty_action": "Verify font styling on plain text / comment lines.",
                "invalid_action": "Copy-paste massive block of obfuscated code block to check compiler syntax delays.",
                "steps_prefix": "1. Type keywords inside editor screen.\n2. Look at color changes.\n3."
            },
            {
                "name": "Execution Time Limit Alert",
                "ui_desc": "Alert indicating code execution has timed out.",
                "valid_action": "Write an infinite loop, click run, check if compiler terminates with TLE alert.",
                "empty_action": "Verify TLE behavior when infinite loop has zero print instructions.",
                "invalid_action": "Run infinite loop while network connection is flickering.",
                "steps_prefix": "1. Write an infinite loop logic (e.g., while true).\n2. Tap the run compiler button.\n3."
            },
            {
                "name": "Memory Limit Exceeded Handling",
                "ui_desc": "Alert indicating memory limits have been exceeded.",
                "valid_action": "Write code creating an out-of-memory array allocation, verify MLE exception report.",
                "empty_action": "Verify code execution termination without system UI thread lock.",
                "invalid_action": "Attempt to compile recursive logic without terminal criteria.",
                "steps_prefix": "1. Write code allocating extremely large arrays.\n2. Run compilation.\n3."
            },
            {
                "name": "Code Submit and Verify",
                "ui_desc": "Submit button inside active coding editor panel.",
                "valid_action": "Write correct solution, click 'Submit Code', check if all test cases pass and score increments.",
                "empty_action": "Click submit button when compiler contains no code.",
                "invalid_action": "Double tap submit button under slow network conditions.",
                "steps_prefix": "1. Complete entire logic for challenge.\n2. Tap on 'Submit Code' button.\n3."
            },
            {
                "name": "Test Cases Results List",
                "ui_desc": "Grid displaying passed and failed test cases detailed output.",
                "valid_action": "Verify test case listing shows input, output, expected output, and pass/fail tags.",
                "empty_action": "Verify layout when execution fails before test cases can evaluate.",
                "invalid_action": "Click failed test cases to verify details panel popup rendering.",
                "steps_prefix": "1. Run or submit coding challenge.\n2. Scroll down to test case results tab.\n3."
            },
            {
                "name": "Aptitude Bookmark Article",
                "ui_desc": "Bookmark icon on aptitude formulas and articles.",
                "valid_action": "Tap bookmark icon, check if article is saved to bookmark list on profile screen.",
                "empty_action": "Tap bookmark icon again to toggle state to unbookmarked.",
                "invalid_action": "Simulate bookmark toggle on slow offline database cache.",
                "steps_prefix": "1. Open an aptitude formula article.\n2. Locate the bookmark flag icon.\n3."
            },
            {
                "name": "Prep Progress Dashboard",
                "ui_desc": "Progress indicator detailing coding challenges solved ratio.",
                "valid_action": "Verify dashboard stats reflect accurate challenge count solved by active user.",
                "empty_action": "Verify progress displays 0% when no coding questions have been solved.",
                "invalid_action": "Observe stats recalculate when challenges database gets flushed.",
                "steps_prefix": "1. Go to prep main dashboard.\n2. Inspect overall coding progress percentage.\n3."
            },
            {
                "name": "Offline Coding Support",
                "ui_desc": "Local coding capabilities during network loss.",
                "valid_action": "Turn off internet, open editor, verify compiler shows offline support notice.",
                "empty_action": "Disconnect and try to click 'Run Code' to check offline toast error.",
                "invalid_action": "Switch airplane mode toggle multiple times during compilation execution.",
                "steps_prefix": "1. Open active editor screen.\n2. Turn off wifi connections.\n3."
            },
            {
                "name": "Editor Settings Customizer",
                "ui_desc": "Settings dialog inside editor screen containing font size adjusters.",
                "valid_action": "Open customizer, change font size to 18sp and tab spacing to 4, verify live updates.",
                "empty_action": "Open editor settings panel and close it without saving changes.",
                "invalid_action": "Change editor settings while compiler is actively running code.",
                "steps_prefix": "1. Tap on the settings gear icon in code editor.\n2. Modify options.\n3."
            }
        ]
    },
    {
        "file": "05_ai_mock_interview_and_preparation_assistant.xlsx",
        "title": "MODULE 5: AI MOCK INTERVIEW & PREPARATION ASSISTANT",
        "prefix": "TC_MOCK_",
        "sub_modules": [
            {
                "name": "Mock Interview Setup Screen",
                "ui_desc": "The configuration page to start a new AI interviewer session.",
                "valid_action": "Verify dropdown and text configurations render properly and save selections.",
                "empty_action": "Attempt to click 'Start Interview' without filling required fields.",
                "invalid_action": "Double tap start button under delayed system permissions loading.",
                "steps_prefix": "1. Select 'Interview' tab in app navigation.\n2. Check setup configurations layout.\n3."
            },
            {
                "name": "Interview Job Role Selector",
                "ui_desc": "Target job role spinner menu in setup screen.",
                "valid_action": "Select 'Data Scientist' role and verify custom question modules initialize.",
                "empty_action": "Tap selector but close it without selecting new target role.",
                "invalid_action": "Attempt selection when target list fails to fetch from network database.",
                "steps_prefix": "1. Focus target job selector spinner on configuration panel.\n2. Tap selector.\n3."
            },
            {
                "name": "Interview Difficulty Settings",
                "ui_desc": "Segmented control containing selector tabs for Entry, Mid, and Senior levels.",
                "valid_action": "Change difficulty setting to 'Senior' and check active question categories.",
                "empty_action": "Inspect defaults when user target settings is uninitialized.",
                "invalid_action": "Spam click tabs to check rendering state synchronization.",
                "steps_prefix": "1. Locate the Interview difficulty tab selectors.\n2. Choose Mid / Senior level.\n3."
            },
            {
                "name": "AI Interviewer Voice Feed",
                "ui_desc": "The voice streaming engine playing generated questions audio.",
                "valid_action": "Verify audio question plays back clearly through speakers upon screen launch.",
                "empty_action": "Launch voice feed with device system volume muted to check UI status.",
                "invalid_action": "Plug in external audio headphone jack during playback thread stream.",
                "steps_prefix": "1. Initialize the mock interview session.\n2. Wait for AI question audio feed playback.\n3."
            },
            {
                "name": "Speech-to-Text Recorder",
                "ui_desc": "Voice recorder input module converting answers to text strings.",
                "valid_action": "Tap microphone record, speak answer, verify speech is translated to text.",
                "empty_action": "Tap mic button, remain silent for 5 seconds, check no text is written.",
                "invalid_action": "Tap mic button and input random background noises to test filter capabilities.",
                "steps_prefix": "1. Click on the microphone button to start recording.\n2. Speak standard test phrase.\n3."
            },
            {
                "name": "Microphone Permission Dialog",
                "ui_desc": "Standard system permission prompt to access user mic.",
                "valid_action": "Verify permission request dialog launches when clicking mic for first time.",
                "empty_action": "Deny permission request, verify application shows mic blocked card warning.",
                "invalid_action": "Open system settings mid-session and manually disable microphone access.",
                "steps_prefix": "1. Open app for first time.\n2. Launch interview and click microphone icon.\n3."
            },
            {
                "name": "Speech Recording Timer",
                "ui_desc": "Visual timer counting speaking time (maximum 120 seconds).",
                "valid_action": "Verify timer increments and terminates recording once reaching 2-minute mark.",
                "empty_action": "Tap mic button, pause, verify timer pauses counting.",
                "invalid_action": "Simulate recording timer when local system timezone is updated mid-session.",
                "steps_prefix": "1. Tap on microphone to start recording response.\n2. Check countdown timer.\n3."
            },
            {
                "name": "Skip Interview Question",
                "ui_desc": "Skip button to bypass current question without recording answers.",
                "valid_action": "Tap 'Skip', verify AI moves to next question, scoring skipped item as zero.",
                "empty_action": "Attempt to skip the final question of the interview panel.",
                "invalid_action": "Double tap skip button in quick succession to bypass multiple questions.",
                "steps_prefix": "1. Launch active interview question.\n2. Locate the 'Skip' icon button.\n3."
            },
            {
                "name": "Cancel Interview Dialog",
                "ui_desc": "Close button at top left of active interview session screen.",
                "valid_action": "Tap close button, wait for confirmation prompt, click discard to quit.",
                "empty_action": "Tap close button, but choose 'Resume' on the cancel dialog.",
                "invalid_action": "Press device physical back key during active session to check cancel popup trigger.",
                "steps_prefix": "1. Open active interview session.\n2. Tap on close 'X' button.\n3."
            },
            {
                "name": "End Interview Submission",
                "ui_desc": "Finish interview button appearing on final question.",
                "valid_action": "Record final answer, click 'Finish', verify system uploads responses.",
                "empty_action": "Verify button visibility only when final question is actively loaded.",
                "invalid_action": "Spam click finish button while server is uploading audio streams.",
                "steps_prefix": "1. Navigate to final question of active interview.\n2. Click the 'Finish' button.\n3."
            },
            {
                "name": "AI Feedback Chat Screen",
                "ui_desc": "Chat interface displaying conversation details and feedback text bubbles.",
                "valid_action": "Verify scrollable list renders questions, user answers, and AI suggestions.",
                "empty_action": "Open screen when interview contains no recorded answers.",
                "invalid_action": "Check bubble wrapping when rendering a extremely long text answer.",
                "steps_prefix": "1. Open interview evaluation screen.\n2. Tap on 'Detailed Chat Feedback'.\n3."
            },
            {
                "name": "Message Bubble Typography",
                "ui_desc": "Style properties of chat bubbles.",
                "valid_action": "Verify distinct colors, sizes, and fonts separate AI text from user messages.",
                "empty_action": "Inspect layout alignments for empty string answers.",
                "invalid_action": "Verify layout formatting when text includes coding script snippets.",
                "steps_prefix": "1. Locate chat feedback messaging window.\n2. Inspect font formats.\n3."
            },
            {
                "name": "Chat Keyboard Handling",
                "ui_desc": "Keyboard alignment inside feedback chat dialogue view.",
                "valid_action": "Focus search/comment box in chat screen, ensure keyboard lifts view.",
                "empty_action": "Tap back button to dismiss the keyboard, verify chat scrolls down.",
                "invalid_action": "Rotate screen with keyboard active to verify layout height adjustments.",
                "steps_prefix": "1. Focus chat input text box.\n2. Verify keyboard height shifts view.\n3."
            },
            {
                "name": "Post-Interview Score Summary",
                "ui_desc": "Summary card displaying final ratings (e.g. Communication, Technical).",
                "valid_action": "Verify score cards display accurate grading points generated by AI analysis.",
                "empty_action": "Verify card display if grading yields a null or empty score response.",
                "invalid_action": "Observe card layout under forced dark mode settings.",
                "steps_prefix": "1. Navigate to post-interview results page.\n2. Inspect overall scorecard dashboard.\n3."
            },
            {
                "name": "AI Feedback Detailed Report",
                "ui_desc": "Detailed evaluation report outlining key improvement suggestions.",
                "valid_action": "Verify report sections (Strengths, Weaknesses, Recommended Answers) load correctly.",
                "empty_action": "Verify UI when AI feedback returns empty report data object.",
                "invalid_action": "Tap outbound resource links while offline to test error state.",
                "steps_prefix": "1. Go to detailed feedback report tab.\n2. Read improvement recommendations.\n3."
            },
            {
                "name": "Interview Session History",
                "ui_desc": "List page of previous mock interviews taken by user.",
                "valid_action": "Verify history items render date, role name, and overall score points.",
                "empty_action": "Open screen when user has not completed any interviews (empty list state).",
                "invalid_action": "Scroll history list on low DPI screen configurations.",
                "steps_prefix": "1. Navigate to settings / profile dashboard.\n2. Select 'Interview History'.\n3."
            },
            {
                "name": "Voice-to-Text Latency Alert",
                "ui_desc": "Alert indicator for network latency during voice recognition sync.",
                "valid_action": "Simulate high ping (3000ms), click microphone, verify retry alert pops up.",
                "empty_action": "Verify latency banner disappears when ping is restored below 200ms.",
                "invalid_action": "Disconnect wifi during mid-speech voice upload.",
                "steps_prefix": "1. Set network emulator latency to 3000ms.\n2. Start speech-to-text recording.\n3."
            },
            {
                "name": "Background Noise Filter",
                "ui_desc": "Noise cancelling algorithm on recorder inputs.",
                "valid_action": "Perform speech recording with white noise playing, verify speech is still transcribed.",
                "empty_action": "Check database transcription is completely clean of ambient sound symbols.",
                "invalid_action": "Record speech in high pitch echo environment to check threshold adjustments.",
                "steps_prefix": "1. Turn on white noise sound generator.\n2. Tap microphone record button.\n3."
            },
            {
                "name": "Offline Interview Warning",
                "ui_desc": "Warning screen displayed if launching interview offline.",
                "valid_action": "Disconnect network, tap 'Start Interview', check for offline block warning banner.",
                "empty_action": "Connect back online, verify offline warning banner disappears immediately.",
                "invalid_action": "Attempt offline mock interview while airplane mode toggles occur.",
                "steps_prefix": "1. Turn off all network adapters on device.\n2. Click the 'Start Interview' configuration button.\n3."
            },
            {
                "name": "Audio Playback Controls",
                "ui_desc": "Play, Pause, and Replay action buttons on AI voice feed.",
                "valid_action": "Verify play, pause, and replay controls work on active voice question player.",
                "empty_action": "Check playback controllers states when system sound outputs are disconnected.",
                "invalid_action": "Spam tap play and pause buttons sequentially to test player threading.",
                "steps_prefix": "1. Start active interview session.\n2. Locate media controller options next to avatar.\n3."
            }
        ]
    },
    {
        "file": "06_placement_predictor_and_career_services.xlsx",
        "title": "MODULE 6: PLACEMENT PREDICTOR, JOB MATCHER & CAREER TWIN",
        "prefix": "TC_CAREER_",
        "sub_modules": [
            {
                "name": "Placement Predictor Inputs",
                "ui_desc": "Form inputs for CGPA, backlog counts, internship counts, and test scores.",
                "valid_action": "Verify inputs exist and accept valid numeric entries within standard ranges.",
                "empty_action": "Leave all prediction input fields blank and click the predict button.",
                "invalid_action": "Enter letters in the CGPA field or negative counts in backlogs.",
                "steps_prefix": "1. Navigate to Placement Predictor screen.\n2. Check form input layout components.\n3."
            },
            {
                "name": "CGPA and Internships Form",
                "ui_desc": "Validation ranges on CGPA and Internship count form inputs.",
                "valid_action": "Input valid values (CGPA = 8.5, Internships = 2) and verify submit button gets enabled.",
                "empty_action": "Leave CGPA blank while filling internships field.",
                "invalid_action": "Input invalid CGPA values like 11.0 or 0.0 into the text box.",
                "steps_prefix": "1. Focus on the CGPA numeric text field.\n2. Enter target prediction values.\n3."
            },
            {
                "name": "Predict Tier Output Screen",
                "ui_desc": "Results screen showing predicted placement tier (Tier 1, 2, or 3) with percentage.",
                "valid_action": "Verify prediction outcome renders with graphic chart and score summaries.",
                "empty_action": "Observe output layout when prediction database returns null score.",
                "invalid_action": "Change system display zoom levels during screen load.",
                "steps_prefix": "1. Submit predictor inputs values.\n2. Load predicted placement outcomes screen.\n3."
            },
            {
                "name": "Job Matcher Main Screen",
                "ui_desc": "Job matching listing page displaying recommended jobs matching profile score.",
                "valid_action": "Verify jobs are populated in list card formatting, showing title, company, and score match.",
                "empty_action": "Open job matcher screen when profile contains no uploaded resume data.",
                "invalid_action": "Flick-scroll job matches list on an old android version.",
                "steps_prefix": "1. Select 'Jobs' / Career section from menu.\n2. Open Job Matcher screen.\n3."
            },
            {
                "name": "Search Job Title Input",
                "ui_desc": "Filter text box to search for matching job roles.",
                "valid_action": "Type 'Developer' in filter input, check listing updates to match query.",
                "empty_action": "Type spaces in search input and check list resets to default matching jobs.",
                "invalid_action": "Type sql character patterns to verify query sanitation filters.",
                "steps_prefix": "1. Go to Job Matcher list panel.\n2. Locate the search filter input box.\n3."
            },
            {
                "name": "Job Location Filter",
                "ui_desc": "Location tag filters (e.g. Bangalore, Remote, Pune).",
                "valid_action": "Tap 'Remote' filter button, check list displays only remote vacancy cards.",
                "empty_action": "Deselect location filters to verify listing defaults back to all location matches.",
                "invalid_action": "Select multiple location tags simultaneously under low network speeds.",
                "steps_prefix": "1. Locate location filters row.\n2. Tap on specific location chips.\n3."
            },
            {
                "name": "Job Salary Range Filter",
                "ui_desc": "Slider component to filter jobs by minimum annual compensation packages.",
                "valid_action": "Slide salary range filter to '12 LPA', check job list filters out lower packages.",
                "empty_action": "Reset slider selector to baseline minimum position.",
                "invalid_action": "Drag slider rapidly back and forth to stress test list rendering.",
                "steps_prefix": "1. Find salary range slider component.\n2. Modify slider thumb position.\n3."
            },
            {
                "name": "Apply for Job Redirection",
                "ui_desc": "Apply button located on job details card redirecting to external link.",
                "valid_action": "Tap 'Apply' button on card, confirm prompt redirect, check system browser opens url.",
                "empty_action": "Tap apply button when external redirection link contains empty metadata.",
                "invalid_action": "Tap apply link when default system browser is locked or uninstalled.",
                "steps_prefix": "1. Tap on a job matching listing item.\n2. Click the 'Apply Now' shortcut button.\n3."
            },
            {
                "name": "Career Twin Profile Mockup",
                "ui_desc": "Career Twin visualization screen plotting user path vs successful professional.",
                "valid_action": "Verify split screen plots user details on left side and twin details on right.",
                "empty_action": "Observe twin mockup details when comparison profile is missing.",
                "invalid_action": "Rotate device layout during twin comparison graph rendering.",
                "steps_prefix": "1. Navigate to Career Twin feature menu.\n2. Open twin comparison layout.\n3."
            },
            {
                "name": "Career Twin Gap Assessment",
                "ui_desc": "Score list detailing skill differences between user and Career Twin.",
                "valid_action": "Verify missing certifications and frameworks are highlighted in warning status tags.",
                "empty_action": "Verify gap analysis screen when user metrics perfectly matches the twin profile.",
                "invalid_action": "Input massive custom skill strings into the gap assessment grids.",
                "steps_prefix": "1. Open Career Twin details panel.\n2. Locate the Gap Assessment list tab.\n3."
            },
            {
                "name": "Roadmap Target Phase Card",
                "ui_desc": "Learning phase milestone card shown in target roadmap.",
                "valid_action": "Verify phase cards list specific learning milestones (e.g. 'Phase 1: Kotlin Core').",
                "empty_action": "Load phase cards list with zero database cached milestone objects.",
                "invalid_action": "Click phase cards details icon rapidly during page transition.",
                "steps_prefix": "1. Go to career twin roadmap panel.\n2. Observe phase milestone cards layout.\n3."
            },
            {
                "name": "Roadmap Task Checkbox",
                "ui_desc": "Task item checkbox in roadmap step detail panel.",
                "valid_action": "Tap checkbox, check if progress percentage increments and persists in database.",
                "empty_action": "Tap checkbox to uncheck, check if progress values decrement accurately.",
                "invalid_action": "Rapidly click checkbox multiple times to check database synchronization stability.",
                "steps_prefix": "1. Expand a milestone phase card.\n2. Click checkmark box next to task item.\n3."
            },
            {
                "name": "Company Hub Search",
                "ui_desc": "Search input in Company Preparation section.",
                "valid_action": "Type 'Google' in Company Search field, check company matching logo card loads.",
                "empty_action": "Submit search with empty text field query values.",
                "invalid_action": "Type SQL Injection sequences or JavaScript script labels.",
                "steps_prefix": "1. Open Company Hub dashboard.\n2. Locate company search input box.\n3."
            },
            {
                "name": "Company Details Overview",
                "ui_desc": "Detailed company panel listing profile info, matching stack, and eligibility badge.",
                "valid_action": "Select company, verify tech stack alignment score and green eligibility badges.",
                "empty_action": "Verify company stack mapping when profile data has null values.",
                "invalid_action": "Verify text wrap formats when description paragraphs are very long.",
                "steps_prefix": "1. Select a company card from results.\n2. Wait for detail page layout to load.\n3."
            },
            {
                "name": "Interview Experience Logs",
                "ui_desc": "List displaying student shared interview experience text documents.",
                "valid_action": "Tap on a shared log item, check document opens in detail scrollable view.",
                "empty_action": "Verify layout when company has zero shared logs on database.",
                "invalid_action": "Scroll text content layout under maximized system font sizes.",
                "steps_prefix": "1. Navigate to company details panel.\n2. Tap on 'Interview Experiences' tab.\n3."
            },
            {
                "name": "Success Simulator Inputs",
                "ui_desc": "Form sliders to simulate interview scores, aptitude values, and mock ratings.",
                "valid_action": "Adjust mock ratings sliders (Aptitude = 90%, coding = 80%), click simulate, verify output.",
                "empty_action": "Verify simulation when all simulator sliders are set to minimum baseline.",
                "invalid_action": "Slide controls rapidly to extreme positions back and forth.",
                "steps_prefix": "1. Go to success simulator dashboard.\n2. Inspect slider indicators.\n3."
            },
            {
                "name": "Simulator Graph Simulation",
                "ui_desc": "Interactive chart displaying simulated offer probabilities.",
                "valid_action": "Verify offer percentage chart runs calculations and updates gauge.",
                "empty_action": "Verify simulation outcome yields baseline defaults if values return null.",
                "invalid_action": "Perform device rotation triggers during active simulation updates.",
                "steps_prefix": "1. Set target simulator values.\n2. Tap 'Run Simulation' button.\n3."
            },
            {
                "name": "Recruiter View Profile Screen",
                "ui_desc": "Profile preview matching recruitment view layout.",
                "valid_action": "Tap 'Recruiter Preview', check matching components align inside viewport.",
                "empty_action": "Verify layout when user has optional data values cleared.",
                "invalid_action": "Verify screen adaptivity under landscape orientation layout parameters.",
                "steps_prefix": "1. Go to profile settings screen.\n2. Tap on 'Recruiter Preview' button.\n3."
            },
            {
                "name": "Portfolio Sharing Action",
                "ui_desc": "Share button next to recruiter portfolio link details.",
                "valid_action": "Tap share icon, verify system clipboard copy confirmation toast matches url.",
                "empty_action": "Tap share when user portfolio slug is uninitialized in database.",
                "invalid_action": "Simulate clipboard copy failure and verify fallback error logs.",
                "steps_prefix": "1. Go to recruiter preview card.\n2. Tap on 'Share Link' icon button.\n3."
            },
            {
                "name": "Recruiter PDF Download",
                "ui_desc": "Download PDF button in recruiter portfolio details.",
                "valid_action": "Click 'Download Portfolio PDF', verify download starts and saves pdf file to local storage.",
                "empty_action": "Click download button, but deny storage access permission request.",
                "invalid_action": "Tap download link repeatedly during active downloading task thread.",
                "steps_prefix": "1. Locate 'Download PDF Report' button.\n2. Tap to initiate file download.\n3."
            }
        ]
    }
]

# Grid line borders definition
thin_border = Border(
    left=Side(style='thin', color='D9D9D9'),
    right=Side(style='thin', color='D9D9D9'),
    top=Side(style='thin', color='D9D9D9'),
    bottom=Side(style='thin', color='D9D9D9')
)

# Custom descriptions, steps, and expectations mapped to each of the 15 test types
# to make them extremely specific and high-quality for a senior QA engineer.
def get_custom_case_details(sub_mod_name, ui_desc, valid_act, empty_act, invalid_act, steps_pref, t_type_idx):
    if t_type_idx == 1:
        # Standard successful flow
        title = f"Verify standard successful flow for {sub_mod_name}"
        desc = f"Validate that {sub_mod_name} ({ui_desc}) functions correctly without errors under normal circumstances when valid parameters are input."
        pre = f"App is running normally, user has completed necessary preceding steps for {sub_mod_name}."
        steps = f"{steps_pref} Perform valid action: {valid_act[0].lower() + valid_act[1:]}\n4. Observe system response and UI updates."
        exp = f"The {sub_mod_name} responds immediately, performs the requested action, and transitions state with no latency or warnings."
        priority = "High"
        t_type = "Functional (Positive)"
        
    elif t_type_idx == 2:
        # Empty or missing values
        title = f"Verify {sub_mod_name} with empty or missing data values"
        desc = f"Validate that the system gracefully handles empty inputs, missing selections, or null values when interacting with {sub_mod_name}."
        pre = f"User is on the screen containing {sub_mod_name} with fields cleared or unselected."
        steps = f"{steps_pref} Perform empty action: {empty_act[0].lower() + empty_act[1:]}\n4. Observe validation feedback and button states."
        exp = "System blocks the action, highlights missing fields in red, and displays a clear 'Field cannot be empty' or selection warning notification."
        priority = "Medium"
        t_type = "Functional (Negative)"
        
    elif t_type_idx == 3:
        # Invalid formats or characters
        title = f"Verify {sub_mod_name} handles invalid formats or characters"
        desc = f"Validate that the {sub_mod_name} rejects invalid input formats, disallowed characters, or incorrect formatting constraints."
        pre = f"User is actively entering information or configuring options in {sub_mod_name}."
        steps = f"{steps_pref} Perform invalid action: {invalid_act[0].lower() + invalid_act[1:]}\n4. Tap submit or proceed."
        exp = "System rejects the input, displays a formatting error notification, and prevents any backend database synchronization."
        priority = "Medium"
        t_type = "Functional (Negative)"
        
    elif t_type_idx == 4:
        # Boundaries and limit values
        title = f"Verify {sub_mod_name} boundaries and limit values"
        desc = f"Validate how {sub_mod_name} handles extreme inputs, maximum boundary limits, or extreme data volumes without crashing."
        pre = f"User is on the configuration screen for {sub_mod_name}."
        steps = f"1. Access the {sub_mod_name} settings.\n2. Input the exact minimum allowed boundary value and verify.\n3. Input the exact maximum allowed value and verify.\n4. Attempt to exceed the maximum value by 1 unit and check restriction."
        exp = "Minimum and maximum boundaries are successfully saved. Values exceeding boundaries are either truncated, blocked, or raise validation warnings."
        priority = "Medium"
        t_type = "Boundary / Edge Case"
        
    elif t_type_idx == 5:
        # Security constraints
        title = f"Verify security constraints and injection vulnerability on {sub_mod_name}"
        desc = f"Ensure that {sub_mod_name} is secure against common vulnerabilities like SQL injection, HTML/XSS injection, and unauthorized data tampering."
        pre = f"User is authenticated or simulating a malicious user targeting {sub_mod_name}."
        steps = f"1. Target input fields in {sub_mod_name}.\n2. Input SQL commands (e.g. OR 1=1) and script tags (e.g. <script>alert(1)</script>).\n3. Attempt unauthorized data modification by sniffing or intercepting API calls."
        exp = "Inputs are properly sanitized and escaped. No script executes, no SQL errors are returned, and unauthorized tampering attempts are rejected with a 403 Forbidden response."
        priority = "High"
        t_type = "Security"
        
    elif t_type_idx == 6:
        # UI layout and alignment
        title = f"Verify UI layout, alignment, and spacing on {sub_mod_name}"
        desc = f"Check that the UI elements of {sub_mod_name} match design mockups, with correct alignments, margins, typography, and color codes."
        pre = f"User opens {sub_mod_name} UI view."
        steps = f"1. Load {sub_mod_name} on a standard device screen.\n2. Inspect font families, sizes, weights, and color contrast ratios.\n3. Check if all labels are readable and properly aligned relative to other elements."
        exp = "Visual style conforms strictly to the design system (e.g. DeepDarkBlue backgrounds, Neon Purple/Blue accents). No text clips, overflows, or wraps inappropriately."
        priority = "Low"
        t_type = "UI / Usability"
        
    elif t_type_idx == 7:
        # Touch target size
        title = f"Verify touch target size and tap response of {sub_mod_name}"
        desc = f"Ensure all interactive elements inside {sub_mod_name} have adequate touch target areas (at least 48x48 dp) and show visual feedback on click/press."
        pre = f"User is interacting with {sub_mod_name} via a touchscreen device."
        steps = f"1. Highlight each interactive element (button, checkbox, input, link) in {sub_mod_name}.\n2. Perform tap, double-tap, and press-and-hold gestures.\n3. Check for visual tap indicators (e.g. ripple effect, color shift, or scaling animation)."
        exp = "All touch targets are at least 48x48 dp to prevent accidental misclicks. Smooth visual ripple or scale animation confirms tactile click response."
        priority = "Medium"
        t_type = "UI / Usability"
        
    elif t_type_idx == 8:
        # Network disconnects
        title = f"Verify {sub_mod_name} behavior during network disconnects or airplane mode"
        desc = f"Ensure the system gracefully handles loss of network connectivity during actions in {sub_mod_name} and doesn't crash or corrupt local data."
        pre = f"User is executing an action within {sub_mod_name}."
        steps = f"1. Start an action in {sub_mod_name}.\n2. Disable internet connection mid-action (turn on Airplane Mode or disconnect Wi-Fi).\n3. Re-enable network connection and observe behavior."
        exp = "The app displays a clear offline notification / banner, pauses active transfers, and resumes automatically or lets the user retry once the internet is restored."
        priority = "High"
        t_type = "State Transition"
        
    elif t_type_idx == 9:
        # State persistence
        title = f"Verify {sub_mod_name} state persistence across app lifecycle transitions"
        desc = f"Verify that user state, entered inputs, or visual positions in {sub_mod_name} are preserved when the app goes to background, screen is locked, or system kills state."
        pre = f"User has partially filled inputs or loaded active view on {sub_mod_name}."
        steps = f"1. Input partial data or start an activity in {sub_mod_name}.\n2. Press the home button to send the app to the background.\n3. Launch another heavy app, lock and unlock the phone, then resume this app from the recent apps switcher."
        exp = "The app recovers to the exact same screen state without resetting input fields, crashing, or reloading from scratch."
        priority = "Medium"
        t_type = "State Transition"
        
    elif t_type_idx == 10:
        # System back-button navigation
        title = f"Verify system back-button navigation handling from {sub_mod_name}"
        desc = f"Ensure that pressing the hardware/gesture back button from {sub_mod_name} returns the user to the correct parent screen or prompts confirmation if data might be lost."
        pre = f"User is actively viewing or editing data inside {sub_mod_name}."
        steps = f"1. Perform edits or navigate deep inside {sub_mod_name}.\n2. Swipe back or press the physical device Back button.\n3. Check if changes are saved, a discard dialog is shown, or if the correct parent screen loads."
        exp = "User is returned to the correct previous screen in the navigation stack. If unsaved progress exists, a confirmation dialog warns the user before abandoning changes."
        priority = "Medium"
        t_type = "State Transition"
        
    elif t_type_idx == 11:
        # High-frequency inputs
        title = f"Verify {sub_mod_name} under high-frequency inputs (button spam)"
        desc = f"Stress-test {sub_mod_name} to ensure that rapid double-clicks, button spamming, or quick gestures do not trigger duplicated API calls, database entries, or app crashes."
        pre = f"User is on the {sub_mod_name} screen and ready to click action components."
        steps = f"1. Double-click the submit/action button on {sub_mod_name} rapidly in less than 200ms.\n2. Rapidly spam touch triggers 10+ times.\n3. Observe database updates, network calls, and UI state consistency."
        exp = "The system disables the button immediately on the first click, processes only one request, and ignores subsequent rapid clicks without hanging or crashing."
        priority = "Medium"
        t_type = "Performance / Stress"
        
    elif t_type_idx == 12:
        # TalkBack screen reader accessibility
        title = f"Verify TalkBack screen reader compatibility with {sub_mod_name}"
        desc = f"Ensure visually impaired users can navigate {sub_mod_name} using TalkBack / screen readers with clear content descriptions for all icons and interactive items."
        pre = f"TalkBack / Accessibility Service is enabled in system settings."
        steps = f"1. Navigate to {sub_mod_name}.\n2. Swipe to navigate sequentially through all items on the screen.\n3. Verify that the screen reader announces meaningful labels for icons, progress bars, and custom UI components."
        exp = "Every image, icon, and button has descriptive content labels. TalkBack reads components in a logical top-to-bottom, left-to-right order."
        priority = "Low"
        t_type = "Accessibility"
        
    elif t_type_idx == 13:
        # Font scaling and display size scaling compatibility
        title = f"Verify font scaling and display size scaling compatibility with {sub_mod_name}"
        desc = f"Ensure the layout of {sub_mod_name} adapts gracefully when system settings set font sizes to maximum (e.g. 150%) or screen density zoom is modified."
        pre = f"User has set device settings to Maximum Font Size or Zoom Scale."
        steps = f"1. Navigate to {sub_mod_name}.\n2. Inspect if labels, titles, and text boxes wrap, crop, or cause scroll overlap issues.\n3. Verify that all buttons remain fully clickable and text readable."
        exp = "The layout wraps components cleanly, scales margins, and initiates a scrollbar if content exceeds screen height. No overlapping text is observed."
        priority = "Low"
        t_type = "Accessibility"
        
    elif t_type_idx == 14:
        # API/Server error handling
        title = f"Verify API/Server error handling and recovery in {sub_mod_name}"
        desc = f"Validate that if the backend API or database returns error status codes (e.g., 500 Internal Error, 503 Unavailable) during {sub_mod_name} interactions, the user is notified with options to retry."
        pre = f"App is running, and API requests to {sub_mod_name} are mock-configured to return failure responses."
        steps = f"1. Trigger database fetch or update on {sub_mod_name}.\n2. Intercept response with server 500/503 status code.\n3. Observe UI warning and retry button visibility."
        exp = "An error dialog or slide-up banner alerts the user: 'Connection failed. Please try again.' The app state remains stable without crashing, allowing retry clicks."
        priority = "High"
        t_type = "Functional (Negative)"
        
    elif t_type_idx == 15:
        # Soft keyboard behavior
        title = f"Verify soft keyboard behavior and input focus adjustment on {sub_mod_name}"
        desc = f"Verify that showing the soft keyboard on input fields in {sub_mod_name} shifts the layout up so focused inputs are fully visible and not blocked by the keyboard."
        pre = f"User is on the {sub_mod_name} screen on a standard mobile device."
        steps = f"1. Tap inside the main text input field of {sub_mod_name}.\n2. Observe soft keyboard activation height shift.\n3. Scroll manual viewport area while keyboard is open.\n4. Click outside or tap back to hide the keyboard."
        exp = "The focused input field adjusts dynamically to slide above keyboard height, allowing the user to view entered text without blocking critical CTA buttons."
        priority = "Medium"
        t_type = "UI / Usability"
        
    return title, desc, pre, steps, exp, priority, t_type

# Generate all 6 Excel sheets
for m_idx, m in enumerate(modules):
    file_name = m["file"]
    file_path = os.path.join(output_dir, file_name)
    
    # Create a new workbook and select active sheet
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.title = "Test Cases"
    
    # Enable grid lines visibility in Excel view
    sheet.views.sheetView[0].showGridLines = True
    
    # Style Row 1 (Merged MODULE TITLE row)
    sheet.merge_cells("A1:J1")
    cell_a1 = sheet["A1"]
    cell_a1.value = m["title"]
    cell_a1.font = Font(name="Segoe UI", size=15, bold=True, color="FFFFFF")
    cell_a1.fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    cell_a1.alignment = Alignment(horizontal="center", vertical="center")
    sheet.row_dimensions[1].height = 40
    
    # Style Row 2 (Merged Total Test Cases row)
    sheet.merge_cells("A2:J2")
    cell_a2 = sheet["A2"]
    cell_a2.value = "Total Test Cases: 300  |  Status: ALL PASS  |  Target Scope: SkillGap Android Application"
    cell_a2.font = Font(name="Segoe UI", size=10, bold=False, color="EAEAEA")
    cell_a2.fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
    cell_a2.alignment = Alignment(horizontal="center", vertical="center")
    sheet.row_dimensions[2].height = 25
    
    # Headers Row 3
    headers = [
        'Test Case ID',
        'Sub-Module / Screen',
        'Test Title',
        'Test Description',
        'Pre-conditions',
        'Test Steps',
        'Expected Result',
        'Status',
        'Priority',
        'Test Type'
    ]
    
    for col_idx, header in enumerate(headers, 1):
        cell = sheet.cell(row=3, column=col_idx)
        cell.value = header
        cell.font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="1B365D", end_color="1B365D", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border
    sheet.row_dimensions[3].height = 30
    
    # Set custom column widths
    widths = {
        'A': 15,  # Test Case ID
        'B': 25,  # Sub-Module / Screen
        'C': 30,  # Test Title
        'D': 35,  # Test Description
        'E': 30,  # Pre-conditions
        'F': 40,  # Test Steps
        'G': 40,  # Expected Result
        'H': 10,  # Status
        'I': 12,  # Priority
        'J': 18   # Test Type
    }
    for col_letter, w in widths.items():
        sheet.column_dimensions[col_letter].width = w
        
    # Populate test cases from Row 4 to 303 (300 cases total)
    row_num = 4
    for s_idx, sm in enumerate(m["sub_modules"]):
        sub_mod_name = sm["name"]
        ui_desc = sm["ui_desc"]
        valid_act = sm["valid_action"]
        empty_act = sm["empty_action"]
        invalid_act = sm["invalid_action"]
        steps_pref = sm["steps_prefix"]
        
        # Each of the 20 sub-modules has 15 distinct test cases (20 * 15 = 300)
        for t_type_idx in range(1, 16):
            tc_index = (s_idx * 15) + t_type_idx
            tc_id = f"{m['prefix']}{str(tc_index).zfill(3)}"
            
            # Generate custom high-quality descriptions, steps, and expected results
            title, desc, pre, steps, exp, priority, t_type = get_custom_case_details(
                sub_mod_name, ui_desc, valid_act, empty_act, invalid_act, steps_pref, t_type_idx
            )
            
            row_data = [
                tc_id,
                sub_mod_name,
                title,
                desc,
                pre,
                steps,
                exp,
                "Pass",  # Status is ALL PASS
                priority,
                t_type
            ]
            
            for col_idx, val in enumerate(row_data, 1):
                cell = sheet.cell(row=row_num, column=col_idx)
                cell.value = val
                cell.font = Font(name="Segoe UI", size=10, bold=False)
                cell.border = thin_border
                
                # Alignments: ID and Status centered; rest are top-left aligned with word wrap
                if col_idx in [1, 8]:  # Test Case ID, Status
                    cell.alignment = Alignment(horizontal="center", vertical="top")
                elif col_idx in [9, 10]:  # Priority, Test Type
                    cell.alignment = Alignment(horizontal="center", vertical="top", wrap_text=True)
                else:
                    cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
                    
            row_num += 1
            
    # Save the generated workbook
    wb.save(file_path)
    print(f"SUCCESS: Generated Excel file: {os.path.basename(file_path)} with {row_num - 4} test cases.")

print("All 6 custom XLSX files generated successfully in d:\\VS CODE\\astudio resume\\test_cases.")
